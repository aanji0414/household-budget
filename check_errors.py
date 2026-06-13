# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\코콕콩\Desktop\annie_claude\가계부_2인가족.xlsx')
errors = []
for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value) if cell.value else ''
            if any(e in v for e in ['#REF!','#DIV/0!','#VALUE!','#NAME?','#N/A']):
                errors.append(f'{sname}!{cell.coordinate}: {v}')

print('Sheets:', wb.sheetnames)
if errors:
    print('ERRORS:', errors)
else:
    print('No formula errors found')
