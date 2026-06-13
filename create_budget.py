from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ── Color constants ──
DARK_BLUE   = "1F3864"
MID_BLUE    = "2F5496"
LIGHT_BLUE  = "BDD7EE"
VLIGHT_BLUE = "DEEAF1"
LIGHT_GREEN = "E2EFDA"
MID_GREEN   = "70AD47"
DARK_GREEN  = "375623"
LIGHT_YELLOW= "FFF2CC"
LIGHT_RED   = "FCE4D6"
PURPLE      = "7030A0"
DARK_PURPLE = "4B0082"
ORANGE      = "C55A11"

BLUE_TEXT   = "0000FF"
BLACK_TEXT  = "000000"
GREEN_TEXT  = "008000"
GRAY_TEXT   = "595959"

KRW = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
NUM = '#,##0;(#,##0);"-"'

def thin_border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def med_border():
    s = Side(style='medium', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=11, color="FFFFFF"):
    return Font(name='Arial', size=size, bold=True, color=color)

def nfont(size=10, bold=False, color=BLACK_TEXT):
    return Font(name='Arial', size=size, bold=bold, color=color)

def ifont(size=10):
    return Font(name='Arial', size=size, bold=False, color=BLUE_TEXT)

def lfont(size=10, bold=False):
    return Font(name='Arial', size=size, bold=bold, color=GREEN_TEXT)

def mk_fill(c):
    return PatternFill("solid", start_color=c, end_color=c)

C = lambda: Alignment(horizontal='center', vertical='center', wrap_text=True)
R = lambda: Alignment(horizontal='right',  vertical='center')
L = lambda: Alignment(horizontal='left',   vertical='center')

def set_cell(ws, addr, value, font=None, fill=None, align=None, fmt=None, border=True):
    cell = ws[addr]
    cell.value = value
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if fmt:    cell.number_format = fmt
    if border: cell.border = thin_border()

def header_row(ws, row, cols, labels, bg, fg="FFFFFF", height=26):
    ws.row_dimensions[row].height = height
    for col, label in zip(cols, labels):
        addr = f'{col}{row}'
        ws[addr] = label
        ws[addr].font = hfont(size=10, color=fg)
        ws[addr].fill = mk_fill(bg)
        ws[addr].alignment = C()
        ws[addr].border = thin_border()

def section_title(ws, row, start_col, end_col, title, bg, fg="FFFFFF", height=26):
    ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
    ws[f'{start_col}{row}'] = title
    ws[f'{start_col}{row}'].font = hfont(color=fg)
    ws[f'{start_col}{row}'].fill = mk_fill(bg)
    ws[f'{start_col}{row}'].alignment = C()
    ws[f'{start_col}{row}'].border = thin_border()
    ws.row_dimensions[row].height = height
    for col_letter in range(ord(start_col)+1, ord(end_col)+1):
        addr = f'{chr(col_letter)}{row}'
        ws[addr].fill = mk_fill(bg)
        ws[addr].border = thin_border()

def page_title(ws, row, start_col, end_col, title, bg, height=40):
    ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
    ws[f'{start_col}{row}'] = title
    ws[f'{start_col}{row}'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    ws[f'{start_col}{row}'].fill = mk_fill(bg)
    ws[f'{start_col}{row}'].alignment = C()
    ws[f'{start_col}{row}'].border = med_border()
    ws.row_dimensions[row].height = height
    for col_letter in range(ord(start_col)+1, ord(end_col)+1):
        addr = f'{chr(col_letter)}{row}'
        ws[addr].fill = mk_fill(bg)

def gap(ws, row, height=8):
    ws.row_dimensions[row].height = height

def data_row(ws, row, col_data, height=22):
    ws.row_dimensions[row].height = height
    for col, (value, font, fill_color, align, fmt) in col_data.items():
        addr = f'{col}{row}'
        ws[addr] = value
        if font:       ws[addr].font = font
        if fill_color: ws[addr].fill = mk_fill(fill_color)
        if align:      ws[addr].alignment = align
        if fmt:        ws[addr].number_format = fmt
        ws[addr].border = thin_border()

# ══════════════════════════════════════════════════
# SHEET 2: 수입
# ══════════════════════════════════════════════════
ws_inc = wb.active
ws_inc.title = "💼 수입"
ws_inc.sheet_view.showGridLines = False
for col, w in [('A',3),('B',28),('C',20),('D',20),('E',20),('F',20),('G',3)]:
    ws_inc.column_dimensions[col].width = w

gap(ws_inc, 1)
page_title(ws_inc, 2, 'B', 'F', "💼 월간 수입 관리", DARK_BLUE)
gap(ws_inc, 3)

# Month row
ws_inc.row_dimensions[4].height = 22
ws_inc['B4'] = "기준 월"
ws_inc['B4'].font = nfont(bold=True)
ws_inc['B4'].fill = mk_fill(LIGHT_BLUE)
ws_inc['B4'].alignment = L()
ws_inc['B4'].border = thin_border()
ws_inc['C4'] = "2026년 06월"
ws_inc['C4'].font = ifont()
ws_inc['C4'].fill = mk_fill(LIGHT_YELLOW)
ws_inc['C4'].alignment = C()
ws_inc['C4'].border = thin_border()
for col in ['D','E','F']:
    ws_inc[f'{col}4'].border = thin_border()

header_row(ws_inc, 5, ['B','C','D','E','F'],
           ['항목', '이번 달 금액 (원)', '전월 금액 (원)', '전년 동월 (원)', '비고'], MID_BLUE)

INC_ROWS = {
    'member1_salary': 6,
    'member2_salary': 7,
    'member1_bonus':  8,
    'member2_bonus':  9,
    'other':          10,
    'total':          11,
}

inc_items = [
    (INC_ROWS['member1_salary'], "구성원 1 월급",         "기본급 + 수당"),
    (INC_ROWS['member2_salary'], "구성원 2 월급",         "기본급 + 수당"),
    (INC_ROWS['member1_bonus'],  "구성원 1 보너스/상여",  "분기별 등"),
    (INC_ROWS['member2_bonus'],  "구성원 2 보너스/상여",  "분기별 등"),
    (INC_ROWS['other'],          "기타 수입 (부업 등)",   ""),
]
for i, (r, label, note) in enumerate(inc_items):
    bg = VLIGHT_BLUE if i % 2 == 0 else "FFFFFF"
    ws_inc.row_dimensions[r].height = 22
    ws_inc[f'B{r}'] = label;  ws_inc[f'B{r}'].font = nfont(); ws_inc[f'B{r}'].fill = mk_fill(bg); ws_inc[f'B{r}'].alignment = L(); ws_inc[f'B{r}'].border = thin_border()
    for col in ['C','D','E']:
        ws_inc[f'{col}{r}'] = 0
        ws_inc[f'{col}{r}'].font = ifont(); ws_inc[f'{col}{r}'].fill = mk_fill(LIGHT_YELLOW)
        ws_inc[f'{col}{r}'].alignment = R(); ws_inc[f'{col}{r}'].number_format = KRW; ws_inc[f'{col}{r}'].border = thin_border()
    ws_inc[f'F{r}'] = note; ws_inc[f'F{r}'].font = nfont(color=GRAY_TEXT); ws_inc[f'F{r}'].fill = mk_fill(bg); ws_inc[f'F{r}'].alignment = L(); ws_inc[f'F{r}'].border = thin_border()

tr = INC_ROWS['total']
ws_inc.row_dimensions[tr].height = 24
ws_inc[f'B{tr}'] = "💰 수입 합계"; ws_inc[f'B{tr}'].font = nfont(bold=True); ws_inc[f'B{tr}'].fill = mk_fill(LIGHT_BLUE); ws_inc[f'B{tr}'].alignment = L(); ws_inc[f'B{tr}'].border = thin_border()
for col in ['C','D','E']:
    ws_inc[f'{col}{tr}'] = f'=SUM({col}6:{col}{tr-1})'
    ws_inc[f'{col}{tr}'].font = nfont(bold=True); ws_inc[f'{col}{tr}'].fill = mk_fill(LIGHT_BLUE)
    ws_inc[f'{col}{tr}'].alignment = R(); ws_inc[f'{col}{tr}'].number_format = KRW; ws_inc[f'{col}{tr}'].border = thin_border()
ws_inc[f'F{tr}'].fill = mk_fill(LIGHT_BLUE); ws_inc[f'F{tr}'].border = thin_border()

# ══════════════════════════════════════════════════
# SHEET 3: 생활비
# ══════════════════════════════════════════════════
ws_exp = wb.create_sheet("🛒 생활비")
ws_exp.sheet_view.showGridLines = False
for col, w in [('A',3),('B',28),('C',20),('D',20),('E',20),('F',20),('G',3)]:
    ws_exp.column_dimensions[col].width = w

gap(ws_exp, 1)
page_title(ws_exp, 2, 'B', 'F', "🛒 2인 가족 생활비 관리", ORANGE)
gap(ws_exp, 3)

ws_exp.row_dimensions[4].height = 22
ws_exp['B4'] = "기준 월"; ws_exp['B4'].font = nfont(bold=True); ws_exp['B4'].fill = mk_fill(LIGHT_RED); ws_exp['B4'].alignment = L(); ws_exp['B4'].border = thin_border()
ws_exp['C4'] = "='💼 수입'!C4"; ws_exp['C4'].font = lfont(bold=True); ws_exp['C4'].fill = mk_fill(LIGHT_YELLOW); ws_exp['C4'].alignment = C(); ws_exp['C4'].border = thin_border()
for col in ['D','E','F']:
    ws_exp[f'{col}4'].border = thin_border()

header_row(ws_exp, 5, ['B','C','D','E','F'],
           ['생활비 항목', '이번 달 금액 (원)', '전월 금액 (원)', '예산 (원)', '비고'], ORANGE)

EXP_ROWS = {
    '주거비': 6, '식비': 7, '교통비': 8, '통신비': 9,
    '의료': 10,  '여가': 11, '교육': 12, '기타': 13,
    'total': 14,
}
exp_items = [
    (6,  "주거비 (월세/관리비)",   "월세, 관리비"),
    (7,  "식비 (외식 포함)",       "마트, 배달, 외식"),
    (8,  "교통비",                  "대중교통, 주유, 주차"),
    (9,  "통신비 (핸드폰 등)",     "2인 합산"),
    (10, "의료/건강비",             "병원, 약, 헬스"),
    (11, "여가/문화생활비",         "여행, OTT, 취미"),
    (12, "교육비",                  "도서, 강의"),
    (13, "기타 생활비",             ""),
]
for i, (r, label, note) in enumerate(exp_items):
    bg = LIGHT_RED if i % 2 == 0 else "FFFFFF"
    ws_exp.row_dimensions[r].height = 22
    ws_exp[f'B{r}'] = label; ws_exp[f'B{r}'].font = nfont(); ws_exp[f'B{r}'].fill = mk_fill(bg); ws_exp[f'B{r}'].alignment = L(); ws_exp[f'B{r}'].border = thin_border()
    for col in ['C','D','E']:
        ws_exp[f'{col}{r}'] = 0
        ws_exp[f'{col}{r}'].font = ifont(); ws_exp[f'{col}{r}'].fill = mk_fill(LIGHT_YELLOW)
        ws_exp[f'{col}{r}'].alignment = R(); ws_exp[f'{col}{r}'].number_format = KRW; ws_exp[f'{col}{r}'].border = thin_border()
    ws_exp[f'F{r}'] = note; ws_exp[f'F{r}'].font = nfont(color=GRAY_TEXT); ws_exp[f'F{r}'].fill = mk_fill(bg); ws_exp[f'F{r}'].alignment = L(); ws_exp[f'F{r}'].border = thin_border()

etr = EXP_ROWS['total']
ws_exp.row_dimensions[etr].height = 24
ws_exp[f'B{etr}'] = "🛒 지출 합계"; ws_exp[f'B{etr}'].font = nfont(bold=True); ws_exp[f'B{etr}'].fill = mk_fill(LIGHT_RED); ws_exp[f'B{etr}'].alignment = L(); ws_exp[f'B{etr}'].border = thin_border()
for col in ['C','D','E']:
    ws_exp[f'{col}{etr}'] = f'=SUM({col}6:{col}{etr-1})'
    ws_exp[f'{col}{etr}'].font = nfont(bold=True); ws_exp[f'{col}{etr}'].fill = mk_fill(LIGHT_RED)
    ws_exp[f'{col}{etr}'].alignment = R(); ws_exp[f'{col}{etr}'].number_format = KRW; ws_exp[f'{col}{etr}'].border = thin_border()
ws_exp[f'F{etr}'].fill = mk_fill(LIGHT_RED); ws_exp[f'F{etr}'].border = thin_border()

# ── 예산 vs 실적 분석 ──
gap(ws_exp, etr+1)
section_title(ws_exp, etr+2, 'B', 'F', "📊 예산 대비 지출 분석", DARK_BLUE)
header_row(ws_exp, etr+3, ['B','C','D','E','F'],
           ['항목', '이번 달 지출', '예산', '잔액 (예산-지출)', '달성률'], MID_BLUE)
for i, (r_src, label, _) in enumerate(exp_items):
    r_ana = etr + 4 + i
    ws_exp.row_dimensions[r_ana].height = 22
    bg = LIGHT_YELLOW if i % 2 == 0 else "FFFFFF"
    ws_exp[f'B{r_ana}'] = f'=B{r_src}'; ws_exp[f'B{r_ana}'].font = lfont(); ws_exp[f'B{r_ana}'].fill = mk_fill(bg); ws_exp[f'B{r_ana}'].alignment = L(); ws_exp[f'B{r_ana}'].border = thin_border()
    ws_exp[f'C{r_ana}'] = f'=C{r_src}'; ws_exp[f'C{r_ana}'].font = lfont(); ws_exp[f'C{r_ana}'].fill = mk_fill(bg); ws_exp[f'C{r_ana}'].alignment = R(); ws_exp[f'C{r_ana}'].number_format = KRW; ws_exp[f'C{r_ana}'].border = thin_border()
    ws_exp[f'D{r_ana}'] = f'=E{r_src}'; ws_exp[f'D{r_ana}'].font = lfont(); ws_exp[f'D{r_ana}'].fill = mk_fill(bg); ws_exp[f'D{r_ana}'].alignment = R(); ws_exp[f'D{r_ana}'].number_format = KRW; ws_exp[f'D{r_ana}'].border = thin_border()
    ws_exp[f'E{r_ana}'] = f'=D{r_ana}-C{r_ana}'; ws_exp[f'E{r_ana}'].font = nfont(); ws_exp[f'E{r_ana}'].fill = mk_fill(bg); ws_exp[f'E{r_ana}'].alignment = R(); ws_exp[f'E{r_ana}'].number_format = KRW; ws_exp[f'E{r_ana}'].border = thin_border()
    ws_exp[f'F{r_ana}'] = f'=IF(D{r_ana}=0,0,C{r_ana}/D{r_ana})'; ws_exp[f'F{r_ana}'].font = nfont(); ws_exp[f'F{r_ana}'].fill = mk_fill(bg); ws_exp[f'F{r_ana}'].alignment = R(); ws_exp[f'F{r_ana}'].number_format = '0.0%'; ws_exp[f'F{r_ana}'].border = thin_border()

# ══════════════════════════════════════════════════
# SHEET 4: 주식투자
# ══════════════════════════════════════════════════
ws_stk = wb.create_sheet("📈 주식투자")
ws_stk.sheet_view.showGridLines = False
for col, w in [('A',3),('B',22),('C',16),('D',16),('E',16),('F',14),('G',16),('H',3)]:
    ws_stk.column_dimensions[col].width = w

gap(ws_stk, 1)
page_title(ws_stk, 2, 'B', 'G', "📈 주식 투자 관리", DARK_GREEN)
gap(ws_stk, 3)
gap(ws_stk, 4)
header_row(ws_stk, 5, ['B','C','D','E','F','G'],
           ['종목명', '투자 원금 (원)', '현재 평가금액 (원)', '평가손익 (원)', '수익률', '이번 달 실현수익 (원)'], MID_GREEN)

STK_DATA_START = 6
STK_DATA_COUNT = 8
for i in range(STK_DATA_COUNT):
    r = STK_DATA_START + i
    ws_stk.row_dimensions[r].height = 22
    bg = LIGHT_GREEN if i % 2 == 0 else "FFFFFF"
    ws_stk[f'B{r}'] = ""; ws_stk[f'B{r}'].font = ifont(); ws_stk[f'B{r}'].fill = mk_fill(LIGHT_YELLOW); ws_stk[f'B{r}'].alignment = L(); ws_stk[f'B{r}'].border = thin_border()
    ws_stk[f'C{r}'] = 0; ws_stk[f'C{r}'].font = ifont(); ws_stk[f'C{r}'].fill = mk_fill(LIGHT_YELLOW); ws_stk[f'C{r}'].alignment = R(); ws_stk[f'C{r}'].number_format = KRW; ws_stk[f'C{r}'].border = thin_border()
    ws_stk[f'D{r}'] = 0; ws_stk[f'D{r}'].font = ifont(); ws_stk[f'D{r}'].fill = mk_fill(LIGHT_YELLOW); ws_stk[f'D{r}'].alignment = R(); ws_stk[f'D{r}'].number_format = KRW; ws_stk[f'D{r}'].border = thin_border()
    ws_stk[f'E{r}'] = f'=D{r}-C{r}'; ws_stk[f'E{r}'].font = nfont(); ws_stk[f'E{r}'].fill = mk_fill(bg); ws_stk[f'E{r}'].alignment = R(); ws_stk[f'E{r}'].number_format = KRW; ws_stk[f'E{r}'].border = thin_border()
    ws_stk[f'F{r}'] = f'=IF(C{r}=0,0,(D{r}-C{r})/C{r})'; ws_stk[f'F{r}'].font = nfont(); ws_stk[f'F{r}'].fill = mk_fill(bg); ws_stk[f'F{r}'].alignment = R(); ws_stk[f'F{r}'].number_format = '0.0%'; ws_stk[f'F{r}'].border = thin_border()
    ws_stk[f'G{r}'] = 0; ws_stk[f'G{r}'].font = ifont(); ws_stk[f'G{r}'].fill = mk_fill(LIGHT_YELLOW); ws_stk[f'G{r}'].alignment = R(); ws_stk[f'G{r}'].number_format = KRW; ws_stk[f'G{r}'].border = thin_border()

# Summary rows
STK_SUM_START = STK_DATA_START + STK_DATA_COUNT  # row 14
gap(ws_stk, STK_SUM_START)
section_title(ws_stk, STK_SUM_START+1, 'B', 'G', "📊 투자 요약", DARK_GREEN)

# STK summary: principal=row16, current=row17, realized=row18, pct=row19
STK_ROWS = {
    'principal': STK_SUM_START+2,
    'current':   STK_SUM_START+3,
    'realized':  STK_SUM_START+4,
    'pct':       STK_SUM_START+5,
}
stk_summary = [
    (STK_ROWS['principal'], "총 투자 원금",        f"=SUM(C{STK_DATA_START}:C{STK_DATA_START+STK_DATA_COUNT-1})", KRW),
    (STK_ROWS['current'],   "현재 총 평가금액",    f"=SUM(D{STK_DATA_START}:D{STK_DATA_START+STK_DATA_COUNT-1})", KRW),
    (STK_ROWS['realized'],  "이번 달 실현 수익금", f"=SUM(G{STK_DATA_START}:G{STK_DATA_START+STK_DATA_COUNT-1})", KRW),
    (STK_ROWS['pct'],       "총 수익률",            f"=IF(C{STK_ROWS['principal']}=0,0,(C{STK_ROWS['current']}-C{STK_ROWS['principal']})/C{STK_ROWS['principal']})", '0.0%'),
]
for i, (r, label, formula, fmt) in enumerate(stk_summary):
    ws_stk.row_dimensions[r].height = 22
    bg = LIGHT_GREEN if i % 2 == 0 else "FFFFFF"
    ws_stk[f'B{r}'] = label; ws_stk[f'B{r}'].font = nfont(bold=True); ws_stk[f'B{r}'].fill = mk_fill(bg); ws_stk[f'B{r}'].alignment = L(); ws_stk[f'B{r}'].border = thin_border()
    ws_stk.merge_cells(f'C{r}:G{r}')
    ws_stk[f'C{r}'] = formula; ws_stk[f'C{r}'].font = nfont(bold=True); ws_stk[f'C{r}'].fill = mk_fill(bg); ws_stk[f'C{r}'].alignment = R(); ws_stk[f'C{r}'].number_format = fmt; ws_stk[f'C{r}'].border = thin_border()
    for col in ['D','E','F','G']:
        ws_stk[f'{col}{r}'].border = thin_border(); ws_stk[f'{col}{r}'].fill = mk_fill(bg)

# ══════════════════════════════════════════════════
# SHEET 5: 적금/예금
# ══════════════════════════════════════════════════
ws_sav = wb.create_sheet("🏦 적금예금")
ws_sav.sheet_view.showGridLines = False
for col, w in [('A',3),('B',24),('C',18),('D',14),('E',14),('F',18),('G',14),('H',3)]:
    ws_sav.column_dimensions[col].width = w

gap(ws_sav, 1)
page_title(ws_sav, 2, 'B', 'G', "🏦 적금 / 예금 현황 관리", PURPLE)
gap(ws_sav, 3)

# ── 적금 ──
section_title(ws_sav, 4, 'B', 'G', "📌 적금 상품 현황", PURPLE)
header_row(ws_sav, 5, ['B','C','D','E','F','G'],
           ['상품명/은행', '월 납입액 (원)', '금리 (%)', '만기일', '누적 잔액 (원)', '명의'], "9B59B6")

SAV_DATA_START = 6
SAV_DATA_COUNT = 5
for i in range(SAV_DATA_COUNT):
    r = SAV_DATA_START + i
    bg = "EAD1DC" if i % 2 == 0 else "FFFFFF"
    ws_sav.row_dimensions[r].height = 22
    ws_sav[f'B{r}'] = ""; ws_sav[f'B{r}'].font = ifont(); ws_sav[f'B{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'B{r}'].alignment = L(); ws_sav[f'B{r}'].border = thin_border()
    ws_sav[f'C{r}'] = 0; ws_sav[f'C{r}'].font = ifont(); ws_sav[f'C{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'C{r}'].alignment = R(); ws_sav[f'C{r}'].number_format = KRW; ws_sav[f'C{r}'].border = thin_border()
    ws_sav[f'D{r}'] = 0; ws_sav[f'D{r}'].font = ifont(); ws_sav[f'D{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'D{r}'].alignment = R(); ws_sav[f'D{r}'].number_format = '0.00%'; ws_sav[f'D{r}'].border = thin_border()
    ws_sav[f'E{r}'] = ""; ws_sav[f'E{r}'].font = ifont(); ws_sav[f'E{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'E{r}'].alignment = C(); ws_sav[f'E{r}'].border = thin_border()
    ws_sav[f'F{r}'] = 0; ws_sav[f'F{r}'].font = ifont(); ws_sav[f'F{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'F{r}'].alignment = R(); ws_sav[f'F{r}'].number_format = KRW; ws_sav[f'F{r}'].border = thin_border()
    ws_sav[f'G{r}'] = ""; ws_sav[f'G{r}'].font = ifont(); ws_sav[f'G{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'G{r}'].alignment = C(); ws_sav[f'G{r}'].border = thin_border()

SAV_TOTAL_ROW = SAV_DATA_START + SAV_DATA_COUNT  # row 11
ws_sav.row_dimensions[SAV_TOTAL_ROW].height = 22
ws_sav[f'B{SAV_TOTAL_ROW}'] = "적금 소계"; ws_sav[f'B{SAV_TOTAL_ROW}'].font = nfont(bold=True); ws_sav[f'B{SAV_TOTAL_ROW}'].fill = mk_fill("EAD1DC"); ws_sav[f'B{SAV_TOTAL_ROW}'].alignment = L(); ws_sav[f'B{SAV_TOTAL_ROW}'].border = thin_border()
ws_sav[f'C{SAV_TOTAL_ROW}'] = f'=SUM(C{SAV_DATA_START}:C{SAV_TOTAL_ROW-1})'; ws_sav[f'C{SAV_TOTAL_ROW}'].font = nfont(bold=True); ws_sav[f'C{SAV_TOTAL_ROW}'].fill = mk_fill("EAD1DC"); ws_sav[f'C{SAV_TOTAL_ROW}'].alignment = R(); ws_sav[f'C{SAV_TOTAL_ROW}'].number_format = KRW; ws_sav[f'C{SAV_TOTAL_ROW}'].border = thin_border()
ws_sav[f'F{SAV_TOTAL_ROW}'] = f'=SUM(F{SAV_DATA_START}:F{SAV_TOTAL_ROW-1})'; ws_sav[f'F{SAV_TOTAL_ROW}'].font = nfont(bold=True); ws_sav[f'F{SAV_TOTAL_ROW}'].fill = mk_fill("EAD1DC"); ws_sav[f'F{SAV_TOTAL_ROW}'].alignment = R(); ws_sav[f'F{SAV_TOTAL_ROW}'].number_format = KRW; ws_sav[f'F{SAV_TOTAL_ROW}'].border = thin_border()
for col in ['D','E','G']:
    ws_sav[f'{col}{SAV_TOTAL_ROW}'].fill = mk_fill("EAD1DC"); ws_sav[f'{col}{SAV_TOTAL_ROW}'].border = thin_border()

# ── 예금 ──
gap(ws_sav, SAV_TOTAL_ROW+1)
DEP_HDR_ROW = SAV_TOTAL_ROW + 2  # row 13
section_title(ws_sav, DEP_HDR_ROW, 'B', 'G', "🏛 예금 상품 현황", "2E4057")
header_row(ws_sav, DEP_HDR_ROW+1, ['B','C','D','E','F','G'],
           ['상품명/은행', '예치 원금 (원)', '금리 (%)', '만기일', '현재 잔액 (원)', '명의'], "4A7C9E")

DEP_DATA_START = DEP_HDR_ROW + 2  # row 15
DEP_DATA_COUNT = 5
for i in range(DEP_DATA_COUNT):
    r = DEP_DATA_START + i
    bg = VLIGHT_BLUE if i % 2 == 0 else "FFFFFF"
    ws_sav.row_dimensions[r].height = 22
    ws_sav[f'B{r}'] = ""; ws_sav[f'B{r}'].font = ifont(); ws_sav[f'B{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'B{r}'].alignment = L(); ws_sav[f'B{r}'].border = thin_border()
    ws_sav[f'C{r}'] = 0; ws_sav[f'C{r}'].font = ifont(); ws_sav[f'C{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'C{r}'].alignment = R(); ws_sav[f'C{r}'].number_format = KRW; ws_sav[f'C{r}'].border = thin_border()
    ws_sav[f'D{r}'] = 0; ws_sav[f'D{r}'].font = ifont(); ws_sav[f'D{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'D{r}'].alignment = R(); ws_sav[f'D{r}'].number_format = '0.00%'; ws_sav[f'D{r}'].border = thin_border()
    ws_sav[f'E{r}'] = ""; ws_sav[f'E{r}'].font = ifont(); ws_sav[f'E{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'E{r}'].alignment = C(); ws_sav[f'E{r}'].border = thin_border()
    ws_sav[f'F{r}'] = 0; ws_sav[f'F{r}'].font = ifont(); ws_sav[f'F{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'F{r}'].alignment = R(); ws_sav[f'F{r}'].number_format = KRW; ws_sav[f'F{r}'].border = thin_border()
    ws_sav[f'G{r}'] = ""; ws_sav[f'G{r}'].font = ifont(); ws_sav[f'G{r}'].fill = mk_fill(LIGHT_YELLOW); ws_sav[f'G{r}'].alignment = C(); ws_sav[f'G{r}'].border = thin_border()

DEP_TOTAL_ROW = DEP_DATA_START + DEP_DATA_COUNT  # row 20
ws_sav.row_dimensions[DEP_TOTAL_ROW].height = 22
ws_sav[f'B{DEP_TOTAL_ROW}'] = "예금 소계"; ws_sav[f'B{DEP_TOTAL_ROW}'].font = nfont(bold=True); ws_sav[f'B{DEP_TOTAL_ROW}'].fill = mk_fill(LIGHT_BLUE); ws_sav[f'B{DEP_TOTAL_ROW}'].alignment = L(); ws_sav[f'B{DEP_TOTAL_ROW}'].border = thin_border()
ws_sav[f'C{DEP_TOTAL_ROW}'] = f'=SUM(C{DEP_DATA_START}:C{DEP_TOTAL_ROW-1})'; ws_sav[f'C{DEP_TOTAL_ROW}'].font = nfont(bold=True); ws_sav[f'C{DEP_TOTAL_ROW}'].fill = mk_fill(LIGHT_BLUE); ws_sav[f'C{DEP_TOTAL_ROW}'].alignment = R(); ws_sav[f'C{DEP_TOTAL_ROW}'].number_format = KRW; ws_sav[f'C{DEP_TOTAL_ROW}'].border = thin_border()
ws_sav[f'F{DEP_TOTAL_ROW}'] = f'=SUM(F{DEP_DATA_START}:F{DEP_TOTAL_ROW-1})'; ws_sav[f'F{DEP_TOTAL_ROW}'].font = nfont(bold=True); ws_sav[f'F{DEP_TOTAL_ROW}'].fill = mk_fill(LIGHT_BLUE); ws_sav[f'F{DEP_TOTAL_ROW}'].alignment = R(); ws_sav[f'F{DEP_TOTAL_ROW}'].number_format = KRW; ws_sav[f'F{DEP_TOTAL_ROW}'].border = thin_border()
for col in ['D','E','G']:
    ws_sav[f'{col}{DEP_TOTAL_ROW}'].fill = mk_fill(LIGHT_BLUE); ws_sav[f'{col}{DEP_TOTAL_ROW}'].border = thin_border()

# ── 전체 저축 요약 ──
gap(ws_sav, DEP_TOTAL_ROW+1)
GRAND_HDR_ROW = DEP_TOTAL_ROW + 2  # row 22
section_title(ws_sav, GRAND_HDR_ROW, 'B', 'G', "💎 저축 전체 요약", PURPLE)

SAV_GRAND = {
    'sav_count':  GRAND_HDR_ROW+1,
    'dep_count':  GRAND_HDR_ROW+2,
    'monthly':    GRAND_HDR_ROW+3,
    'sav_bal':    GRAND_HDR_ROW+4,
    'dep_bal':    GRAND_HDR_ROW+5,
    'total_bal':  GRAND_HDR_ROW+6,
}
grand_items = [
    (SAV_GRAND['sav_count'],  "적금 상품 수 (개)",       f'=COUNTA(B{SAV_DATA_START}:B{SAV_TOTAL_ROW-1})', '#,##0'),
    (SAV_GRAND['dep_count'],  "예금 상품 수 (개)",       f'=COUNTA(B{DEP_DATA_START}:B{DEP_TOTAL_ROW-1})', '#,##0'),
    (SAV_GRAND['monthly'],    "이번 달 적금 납입 총액",  f'=C{SAV_TOTAL_ROW}', KRW),
    (SAV_GRAND['sav_bal'],    "적금 누적 잔액 합계",     f'=F{SAV_TOTAL_ROW}', KRW),
    (SAV_GRAND['dep_bal'],    "예금 잔액 합계",           f'=F{DEP_TOTAL_ROW}', KRW),
    (SAV_GRAND['total_bal'],  "저축 총 잔액",             f'=F{SAV_TOTAL_ROW}+F{DEP_TOTAL_ROW}', KRW),
]
for i, (r, label, formula, fmt) in enumerate(grand_items):
    ws_sav.row_dimensions[r].height = 22
    bg = "EAD1DC" if i % 2 == 0 else "FFFFFF"
    ws_sav[f'B{r}'] = label; ws_sav[f'B{r}'].font = nfont(bold=True); ws_sav[f'B{r}'].fill = mk_fill(bg); ws_sav[f'B{r}'].alignment = L(); ws_sav[f'B{r}'].border = thin_border()
    ws_sav.merge_cells(f'C{r}:G{r}')
    ws_sav[f'C{r}'] = formula; ws_sav[f'C{r}'].font = nfont(bold=True); ws_sav[f'C{r}'].fill = mk_fill(bg); ws_sav[f'C{r}'].alignment = R(); ws_sav[f'C{r}'].number_format = fmt; ws_sav[f'C{r}'].border = thin_border()
    for col in ['D','E','F','G']:
        ws_sav[f'{col}{r}'].border = thin_border(); ws_sav[f'{col}{r}'].fill = mk_fill(bg)

# ══════════════════════════════════════════════════
# SHEET 1: 대시보드 (built last, now all row refs known)
# ══════════════════════════════════════════════════
ws_dash = wb.create_sheet("📊 대시보드", 0)
ws_dash.sheet_view.showGridLines = False
for col, w in [('A',3),('B',24),('C',20),('D',4),('E',24),('F',20),('G',4)]:
    ws_dash.column_dimensions[col].width = w

gap(ws_dash, 1)
ws_dash.merge_cells('B2:F2')
ws_dash['B2'] = "🏠 2인 가족 가계부 대시보드"
ws_dash['B2'].font = Font(name='Arial', size=18, bold=True, color="FFFFFF")
ws_dash['B2'].fill = mk_fill(DARK_BLUE)
ws_dash['B2'].alignment = C()
ws_dash['B2'].border = med_border()
ws_dash.row_dimensions[2].height = 44
for col in ['C','D','E','F']:
    ws_dash[f'{col}2'].fill = mk_fill(DARK_BLUE)
gap(ws_dash, 3)

# ── 수입 요약 ──
section_title(ws_dash, 4, 'B', 'C', "💰 수입 요약", MID_BLUE)
inc_refs = [
    ("구성원 1 월급",  f"='💼 수입'!C{INC_ROWS['member1_salary']}", False),
    ("구성원 2 월급",  f"='💼 수입'!C{INC_ROWS['member2_salary']}", False),
    ("보너스 합계",    f"='💼 수입'!C{INC_ROWS['member1_bonus']}+'💼 수입'!C{INC_ROWS['member2_bonus']}", False),
    ("기타 수입",      f"='💼 수입'!C{INC_ROWS['other']}", False),
    ("수입 합계",      f"='💼 수입'!C{INC_ROWS['total']}", True),
]
for i, (label, formula, bold) in enumerate(inc_refs):
    r = 5 + i
    bg = VLIGHT_BLUE if i % 2 == 0 else "FFFFFF"
    ws_dash.row_dimensions[r].height = 22
    ws_dash[f'B{r}'] = label; ws_dash[f'B{r}'].font = nfont(bold=bold); ws_dash[f'B{r}'].fill = mk_fill(bg); ws_dash[f'B{r}'].alignment = L(); ws_dash[f'B{r}'].border = thin_border()
    ws_dash[f'C{r}'] = formula; ws_dash[f'C{r}'].font = lfont(bold=bold); ws_dash[f'C{r}'].fill = mk_fill(bg); ws_dash[f'C{r}'].alignment = R(); ws_dash[f'C{r}'].number_format = KRW; ws_dash[f'C{r}'].border = thin_border()

gap(ws_dash, 10)

# ── 지출 요약 ──
section_title(ws_dash, 11, 'B', 'C', "🛒 지출 요약", ORANGE)
exp_refs = [
    ("주거비",          f"='🛒 생활비'!C{EXP_ROWS['주거비']}", False),
    ("식비",            f"='🛒 생활비'!C{EXP_ROWS['식비']}", False),
    ("교통비",          f"='🛒 생활비'!C{EXP_ROWS['교통비']}", False),
    ("통신비",          f"='🛒 생활비'!C{EXP_ROWS['통신비']}", False),
    ("의료/건강",       f"='🛒 생활비'!C{EXP_ROWS['의료']}", False),
    ("여가/문화",       f"='🛒 생활비'!C{EXP_ROWS['여가']}", False),
    ("교육비",          f"='🛒 생활비'!C{EXP_ROWS['교육']}", False),
    ("기타",            f"='🛒 생활비'!C{EXP_ROWS['기타']}", False),
    ("지출 합계",       f"='🛒 생활비'!C{EXP_ROWS['total']}", True),
]
for i, (label, formula, bold) in enumerate(exp_refs):
    r = 12 + i
    bg = LIGHT_RED if i % 2 == 0 else "FFFFFF"
    ws_dash.row_dimensions[r].height = 22
    ws_dash[f'B{r}'] = label; ws_dash[f'B{r}'].font = nfont(bold=bold); ws_dash[f'B{r}'].fill = mk_fill(bg); ws_dash[f'B{r}'].alignment = L(); ws_dash[f'B{r}'].border = thin_border()
    ws_dash[f'C{r}'] = formula; ws_dash[f'C{r}'].font = lfont(bold=bold); ws_dash[f'C{r}'].fill = mk_fill(bg); ws_dash[f'C{r}'].alignment = R(); ws_dash[f'C{r}'].number_format = KRW; ws_dash[f'C{r}'].border = thin_border()

gap(ws_dash, 21)

# ── 월간 잔액 ──
section_title(ws_dash, 22, 'B', 'C', "💡 월간 순잔액", MID_GREEN)
bal_r_inc = INC_ROWS['total']
bal_r_exp = EXP_ROWS['total']
bal_r_sav = SAV_GRAND['monthly']
balance_refs = [
    ("월 수입 합계",        f"='💼 수입'!C{bal_r_inc}", False),
    ("월 지출 합계",        f"='🛒 생활비'!C{bal_r_exp}", False),
    ("적금 납입 총액",      f"='🏦 적금예금'!C{bal_r_sav}", False),
    ("순잔액 (수입-지출-적금)", f"='💼 수입'!C{bal_r_inc}-'🛒 생활비'!C{bal_r_exp}-'🏦 적금예금'!C{bal_r_sav}", True),
]
for i, (label, formula, bold) in enumerate(balance_refs):
    r = 23 + i
    bg = LIGHT_GREEN if i % 2 == 0 else "FFFFFF"
    ws_dash.row_dimensions[r].height = 22
    ws_dash[f'B{r}'] = label; ws_dash[f'B{r}'].font = nfont(bold=bold); ws_dash[f'B{r}'].fill = mk_fill(bg); ws_dash[f'B{r}'].alignment = L(); ws_dash[f'B{r}'].border = thin_border()
    ws_dash[f'C{r}'] = formula; ws_dash[f'C{r}'].font = lfont(bold=bold); ws_dash[f'C{r}'].fill = mk_fill(bg); ws_dash[f'C{r}'].alignment = R(); ws_dash[f'C{r}'].number_format = KRW; ws_dash[f'C{r}'].border = thin_border()

# ── 주식 요약 (우측) ──
section_title(ws_dash, 4, 'E', 'F', "📈 주식 투자 현황", DARK_GREEN)
stk_refs = [
    ("총 투자 원금",      f"='📈 주식투자'!C{STK_ROWS['principal']}", False),
    ("현재 평가금액",     f"='📈 주식투자'!C{STK_ROWS['current']}", False),
    ("이번 달 수익금",    f"='📈 주식투자'!C{STK_ROWS['realized']}", False),
    ("총 수익률",         f"='📈 주식투자'!C{STK_ROWS['pct']}", False),
]
fmt_stk = [KRW, KRW, KRW, '0.0%']
for i, ((label, formula, bold), fmt) in enumerate(zip(stk_refs, fmt_stk)):
    r = 5 + i
    bg = LIGHT_GREEN if i % 2 == 0 else "FFFFFF"
    ws_dash.row_dimensions[r].height = 22
    ws_dash[f'E{r}'] = label; ws_dash[f'E{r}'].font = nfont(bold=bold); ws_dash[f'E{r}'].fill = mk_fill(bg); ws_dash[f'E{r}'].alignment = L(); ws_dash[f'E{r}'].border = thin_border()
    ws_dash[f'F{r}'] = formula; ws_dash[f'F{r}'].font = lfont(bold=bold); ws_dash[f'F{r}'].fill = mk_fill(bg); ws_dash[f'F{r}'].alignment = R(); ws_dash[f'F{r}'].number_format = fmt; ws_dash[f'F{r}'].border = thin_border()

gap(ws_dash, 9)

# ── 저축 요약 (우측) ──
section_title(ws_dash, 10, 'E', 'F', "🏦 적금/예금 현황", PURPLE)
sav_refs = [
    ("적금 상품 수",      f"='🏦 적금예금'!C{SAV_GRAND['sav_count']}", '#,##0'),
    ("예금 상품 수",      f"='🏦 적금예금'!C{SAV_GRAND['dep_count']}", '#,##0'),
    ("월 납입 총액",      f"='🏦 적금예금'!C{SAV_GRAND['monthly']}", KRW),
    ("적금 누적 잔액",    f"='🏦 적금예금'!C{SAV_GRAND['sav_bal']}", KRW),
    ("예금 잔액",         f"='🏦 적금예금'!C{SAV_GRAND['dep_bal']}", KRW),
    ("저축 총 잔액",      f"='🏦 적금예금'!C{SAV_GRAND['total_bal']}", KRW),
]
for i, (label, formula, fmt) in enumerate(sav_refs):
    r = 11 + i
    bg = "EAD1DC" if i % 2 == 0 else "FFFFFF"
    ws_dash.row_dimensions[r].height = 22
    ws_dash[f'E{r}'] = label; ws_dash[f'E{r}'].font = nfont(); ws_dash[f'E{r}'].fill = mk_fill(bg); ws_dash[f'E{r}'].alignment = L(); ws_dash[f'E{r}'].border = thin_border()
    ws_dash[f'F{r}'] = formula; ws_dash[f'F{r}'].font = lfont(); ws_dash[f'F{r}'].fill = mk_fill(bg); ws_dash[f'F{r}'].alignment = R(); ws_dash[f'F{r}'].number_format = fmt; ws_dash[f'F{r}'].border = thin_border()

# ── 범례 ──
legend_r = 27
gap(ws_dash, legend_r)
ws_dash.merge_cells(f'B{legend_r+1}:F{legend_r+1}')
ws_dash[f'B{legend_r+1}'] = "📌 색상 범례:  파란색 글씨 = 직접 입력값   |   초록색 글씨 = 다른 시트 연동값"
ws_dash[f'B{legend_r+1}'].font = Font(name='Arial', size=9, italic=True, color=GRAY_TEXT)
ws_dash[f'B{legend_r+1}'].alignment = C()
ws_dash.row_dimensions[legend_r+1].height = 18

output_path = r"C:\Users\코콕콩\Desktop\annie_claude\가계부_2인가족.xlsx"
wb.save(output_path)
print(f"저장 완료: {output_path}")
